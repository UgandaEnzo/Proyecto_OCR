import os
from pathlib import Path
from typing import Optional, List
import traceback
import sys
base_dir = Path(sys.executable if getattr(sys, 'frozen', False) else __file__).resolve().parent
uploads_dir = base_dir / 'uploads'
from datetime import datetime, timedelta
import io
import re
import json
import base64
from PIL import Image
from sqlalchemy import func, text
from sqlalchemy.orm import Session
from fastapi import Header, HTTPException, Depends
import logging
from logging.handlers import RotatingFileHandler

import models
import bank_rules
from database import SQLALCHEMY_DATABASE_URL, get_db
from config import get_config_value

# Configuración del logger
def _setup_logging() -> logging.Logger:
    os.makedirs("logs", exist_ok=True)
    logger = logging.getLogger("ocr_api")
    if logger.handlers:
        return logger
    logger.setLevel(os.getenv("LOG_LEVEL", "INFO").upper())
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(name)s - %(message)s")
    file_handler = RotatingFileHandler(os.path.join("logs", "app.log"), maxBytes=2 * 1024 * 1024, backupCount=5, encoding="utf-8")
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    console = logging.StreamHandler()
    console.setFormatter(fmt)
    logger.addHandler(console)
    return logger

logger = _setup_logging()

def require_api_key(x_api_key: Optional[str]=Header(None), db: Session = Depends(get_db)):
    configured_key = get_config_value(db, 'API_KEY')
    if configured_key and (not x_api_key or x_api_key != configured_key):
        raise HTTPException(status_code=401, detail='API key inválida o no proporcionada')
    return True

def registrar_auditoria(db: Session, pago_id: int, accion: str, detalles: str):
    nuevo_historial = models.PagoHistory(pago_id=pago_id, accion=accion, detalles=detalles, usuario='sistema_ia')
    db.add(nuevo_historial)
    db.commit()

def set_env_value(key: str, value: str):
    """Escribe o actualiza una variable en el archivo .env."""
    env_path = Path('.env')
    if env_path.exists():
        lines = env_path.read_text(encoding='utf-8').splitlines()
    else:
        lines = []
    key_upper = key.upper()
    found = False
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('#') or '=' not in stripped:
            continue
        k = stripped.split('=', 1)[0].strip().upper()
        if k == key_upper:
            lines[i] = f'{key_upper}={value}'
            found = True
            break
    if not found:
        lines.append(f'{key_upper}={value}')
    env_path.write_text('\n'.join(lines) + '\n', encoding='utf-8')

def _get_database_type() -> str:
    url = SQLALCHEMY_DATABASE_URL.lower()
    if url.startswith('sqlite://'):
        return 'sqlite'
    if url.startswith('postgresql://') or url.startswith('postgres://'):
        return 'postgresql'
    return 'unsupported'

def _get_sqlite_db_path() -> Optional[Path]:
    if _get_database_type() != 'sqlite':
        return None
    if SQLALCHEMY_DATABASE_URL.startswith('sqlite:///'):
        sqlite_path = SQLALCHEMY_DATABASE_URL.replace('sqlite:///', '', 1)
        return Path(sqlite_path).resolve()
    if SQLALCHEMY_DATABASE_URL.startswith('sqlite://'):
        sqlite_path = SQLALCHEMY_DATABASE_URL.replace('sqlite://', '', 1)
        return Path(sqlite_path).resolve()
    return None

async def _verificar_estado_ia(api_key: str) -> tuple[bool, str]:
    if not api_key:
        return (False, 'No se ha configurado la clave OpenRouter.')
    try:
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")
        from ai_client import openrouter
        await client.chat.completions.create(messages=[{'role': 'user', 'content': 'Responde con pong'}], model=openrouter.TEXT_MODEL, temperature=0.0, max_tokens=1)
        return (True, 'Clave OpenRouter verificada correctamente.')
    except Exception as e:
        logger.warning('No se pudo verificar OpenRouter: %s', e)
        return (False, 'No se puede conectar a OpenRouter. Comprueba tu conexion y clave API.')

def _agregar_total_reporte(resultado: List[dict]) -> dict:
    total_bs = sum((item['total_bs'] for item in resultado))
    total_usd = sum((item['total_usd'] for item in resultado))
    total_pagos = sum((item['conteo'] for item in resultado))
    return {'total_bs': total_bs, 'total_usd': total_usd, 'total_pagos': total_pagos}

def _query_reporte(db: Session, tipo_reporte: str, fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime]) -> List[dict]:
    if tipo_reporte == 'general':
        query = db.query(func.sum(models.Pago.monto).label('total_bs'), func.coalesce(func.sum(models.Pago.monto_usd), 0).label('total_usd'), func.count(models.Pago.id).label('conteo'), func.min(models.Pago.fecha_registro).label('desde'), func.max(models.Pago.fecha_registro).label('hasta'))
        if fecha_inicio:
            query = query.filter(models.Pago.fecha_registro >= fecha_inicio)
        if fecha_fin:
            query = query.filter(models.Pago.fecha_registro <= fecha_fin)
        result = query.first()
        if result:
            return [{'periodo': 'General', 'desde': result.desde, 'hasta': result.hasta, 'total_bs': float(result.total_bs or 0), 'total_usd': float(result.total_usd or 0), 'conteo': int(result.conteo or 0)}]
        return []
    interval_map = {'diario': 'day', 'semanal': 'week', 'mensual': 'month', 'trimestral': 'quarter', 'anual': 'year'}
    if tipo_reporte not in interval_map and tipo_reporte not in ['quincenal', 'semestral']:
        raise HTTPException(status_code=400, detail=f'Tipo de reporte desconocido: {tipo_reporte}')
    local_ts = models.Pago.fecha_registro.op('AT TIME ZONE')('America/Caracas')
    if tipo_reporte == 'quincenal':
        periodo_expr = func.date_trunc('month', local_ts) + func.floor((func.extract('day', local_ts) - 1) / 15) * text("interval '15 days'")
    elif tipo_reporte == 'semestral':
        periodo_expr = func.date_trunc('year', local_ts) + func.floor((func.extract('month', local_ts) - 1) / 6) * text("interval '6 months'")
    else:
        periodo_expr = func.date_trunc(interval_map[tipo_reporte], local_ts)
    query = db.query(periodo_expr.label('periodo'), func.sum(models.Pago.monto).label('total_bs'), func.coalesce(func.sum(models.Pago.monto_usd), 0).label('total_usd'), func.count(models.Pago.id).label('conteo'), func.min(models.Pago.fecha_registro).label('desde'), func.max(models.Pago.fecha_registro).label('hasta'))
    if fecha_inicio:
        query = query.filter(models.Pago.fecha_registro >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Pago.fecha_registro <= fecha_fin)
    grouped = query.group_by(periodo_expr).order_by(periodo_expr).all()
    resultados = []
    for r in grouped:
        periodo_text = _simplificar_periodo(r.periodo)
        resultados.append({'periodo': periodo_text, 'desde': r.desde, 'hasta': r.hasta, 'total_bs': float(r.total_bs or 0), 'total_usd': float(r.total_usd or 0), 'conteo': int(r.conteo or 0)})
    return resultados

def _simplificar_periodo(periodo: Optional[object]) -> str:
    if periodo is None:
        return ''
    if isinstance(periodo, datetime):
        return periodo.strftime('%d-%m-%Y')
    if isinstance(periodo, str):
        periodo = periodo.strip()
        if not periodo:
            return ''
        try:
            if 'T' in periodo:
                fecha = datetime.fromisoformat(periodo)
                return fecha.strftime('%d-%m-%Y')
        except Exception:
            pass
        if 'T' in periodo:
            return periodo.split('T')[0]
        if ' ' in periodo and ':' in periodo:
            return periodo.split(' ')[0]
        return periodo
    return str(periodo)

def _limpiar_periodo_texto(periodo: Optional[object]) -> str:
    if periodo is None:
        return ''
    if isinstance(periodo, datetime):
        return periodo.strftime('%d-%m-%Y')
    if isinstance(periodo, str):
        texto = periodo.strip()
        if not texto:
            return ''
        match = re.search('(\\d{2})[-/](\\d{2})[-/](\\d{4})', texto)
        if match:
            return f'{match.group(1)}-{match.group(2)}-{match.group(3)}'
        match = re.search('(\\d{4})-(\\d{2})-(\\d{2})', texto)
        if match:
            return f'{match.group(3)}-{match.group(2)}-{match.group(1)}'
        texto = texto.splitlines()[0].strip()
        return texto if len(texto) <= 32 else texto[:32] + '...'
    return str(periodo)

def _query_pagos_detalle(db: Session, fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime]) -> List[models.Pago]:
    """Obtiene la lista detallada de pagos para el reporte."""
    query = db.query(models.Pago)
    if fecha_inicio:
        query = query.filter(models.Pago.fecha_registro >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Pago.fecha_registro <= fecha_fin)
    return query.order_by(models.Pago.fecha_registro.desc()).all()

def parse_monto_string(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return 0.0
    texto = re.sub('\\s+', '', texto)
    texto = texto.replace('Bs', '').replace('Bs.', '').replace('Bs,', '')
    texto = re.sub('[^0-9\\,\\.\\-]', '', texto)
    if texto.count(',') > 0 and texto.count('.') > 0:
        if texto.rfind(',') > texto.rfind('.'):
            texto = texto.replace('.', '')
            texto = texto.replace(',', '.')
        else:
            texto = texto.replace(',', '')
    else:
        texto = texto.replace(',', '.')
    try:
        return float(texto)
    except ValueError:
        return 0.0

def _extraer_codigo_sudeban(texto: Optional[str]) -> Optional[str]:
    if not texto:
        return None
    texto = str(texto)
    match = re.search('\\b(\\d{4})\\b', texto)
    if match:
        return match.group(1)
    return None

def _agrupar_totales_sudeban(pagos_detalle: List[models.Pago]) -> List[dict]:
    grupos = {}
    for pago in pagos_detalle:
        codigo = _extraer_codigo_sudeban(pago.banco) or _extraer_codigo_sudeban(pago.banco_destino)
        if codigo:
            banco_formal = bank_rules.get_bank_by_sudeban_code(codigo)
            etiqueta = f'{codigo} - {banco_formal}' if banco_formal != 'Desconocido' else f'{codigo} - Desconocido'
        else:
            etiqueta = 'Desconocido'
            codigo = 'N/A'
        clave = (codigo, etiqueta)
        if clave not in grupos:
            grupos[clave] = {'sudeban_code': codigo, 'banco_label': etiqueta, 'total_bs': 0.0, 'total_usd': 0.0, 'conteo': 0}
        grupos[clave]['total_bs'] += parse_monto_string(pago.monto)
        grupos[clave]['total_usd'] += parse_monto_string(pago.monto_usd)
        grupos[clave]['conteo'] += 1
    return [{'sudeban_code': codigo, 'banco_label': etiqueta, 'total_bs': valores['total_bs'], 'total_usd': valores['total_usd'], 'conteo': valores['conteo']} for (codigo, etiqueta), valores in sorted(grupos.items(), key=lambda item: item[0])]

def _crear_excel_reporte(resultados: List[dict], pagos_detalle: List[models.Pago], tipo_reporte: str, start_date: Optional[datetime], end_date: Optional[datetime], empresa_nombre: str = '', color_primario: str = '#1e3a8a', color_secundario: str = '#dbeafe', logo_bytes: Optional[bytes] = None, rif: str = '', contacto: str = '') -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from PIL import Image as PILImage
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    header_fill = PatternFill(start_color=color_primario.lstrip('#'), end_color=color_primario.lstrip('#'), fill_type='solid')
    sub_header_fill = PatternFill(start_color=color_secundario.lstrip('#'), end_color=color_secundario.lstrip('#'), fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='e5e7eb'),
        left=Side(style='thin', color='e5e7eb'),
        right=Side(style='thin', color='e5e7eb')
    )
    right_align = Alignment(horizontal='right', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(wrap_text=True, vertical='top')

    # --- Membrete centrado (independiente de las tablas) ---
    fila = 1
    if logo_bytes:
        try:
            pil_img = PILImage.open(io.BytesIO(logo_bytes))
            ow, oh = pil_img.size
            ratio = min(120/ow, 60/oh)
            xl_img = XlImage(io.BytesIO(logo_bytes))
            xl_img.width = int(ow * ratio)
            xl_img.height = int(oh * ratio)
            ws.add_image(xl_img, f'C{fila}')
            fila += 3
        except Exception:
            pass
    titulo = empresa_nombre + ' - ' if empresa_nombre else ''
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    celda = ws.cell(row=fila, column=1, value=f'{titulo}Reporte de Conciliación Bancaria')
    celda.font = title_font
    celda.alignment = center_align
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
    celda = ws.cell(row=fila, column=1, value=f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}  |  Tipo: {tipo_reporte.title()}')
    celda.font = normal_font
    celda.alignment = center_align
    fila += 1
    texto_periodo = ''
    if start_date:
        texto_periodo += f'Desde: {start_date.strftime("%Y-%m-%d")}'
    if end_date:
        texto_periodo += f'  Hasta: {end_date.strftime("%Y-%m-%d")}' if texto_periodo else f'Hasta: {end_date.strftime("%Y-%m-%d")}'
    if texto_periodo:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        celda = ws.cell(row=fila, column=1, value=texto_periodo)
        celda.font = normal_font
        celda.alignment = center_align
        fila += 1
    if rif:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        celda = ws.cell(row=fila, column=1, value=f'RIF: {rif}')
        celda.font = normal_font
        celda.alignment = center_align
        fila += 1
    if contacto:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=6)
        celda = ws.cell(row=fila, column=1, value=f'Contacto: {contacto}')
        celda.font = normal_font
        celda.alignment = center_align
        fila += 1
    fila += 1

    # --- Resumen Agregado ---
    row_actual = ws.max_row + 1
    ws.append(['Resumen Agregado'])
    ws[row_actual][0].font = section_font
    ws.append(['Periodo', 'Desde', 'Hasta', 'Total Bs', 'Total USD', 'Conteo'])
    row_headers = ws.max_row
    for cell in ws[row_headers]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    for item in resultados:
        periodo_text = _limpiar_periodo_texto(item['periodo'])
        ws.append([periodo_text,
                   item['desde'].strftime('%Y-%m-%d') if item['desde'] else '',
                   item['hasta'].strftime('%Y-%m-%d') if item['hasta'] else '',
                   parse_monto_string(item.get('total_bs')),
                   parse_monto_string(item.get('total_usd')),
                   item.get('conteo', 0)])
        r = ws.max_row
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).font = normal_font
            if c >= 4:
                ws.cell(row=r, column=c).alignment = right_align
                ws.cell(row=r, column=c).number_format = '#,##0.00'
            if c == 6:
                ws.cell(row=r, column=c).alignment = center_align
        ws.cell(row=r, column=1).alignment = wrap_align
    totales = _agregar_total_reporte(resultados)
    ws.append(['Totales', '', '', totales['total_bs'], totales['total_usd'], totales['total_pagos']])
    r = ws.max_row
    for c in range(1, 7):
        ws.cell(row=r, column=c).font = Font(bold=True, size=10)
        ws.cell(row=r, column=c).border = thin_border
        if c >= 4:
            ws.cell(row=r, column=c).alignment = right_align
            ws.cell(row=r, column=c).number_format = '#,##0.00'
        if c == 6:
            ws.cell(row=r, column=c).alignment = center_align
    ws.append([])

    # --- Detalle Individual de Pagos ---
    row_actual = ws.max_row + 1
    ws.append(['Detalle Individual de Pagos'])
    ws[row_actual][0].font = section_font
    ws.append(['Referencia', 'Banco', 'Fecha', 'Monto Bs', 'Tasa ($)', 'Monto USD'])
    row_headers = ws.max_row
    for cell in ws[row_headers]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    if pagos_detalle:
        for p in pagos_detalle:
            ws.append([p.referencia,
                       p.banco or '-',
                       p.fecha_registro.strftime('%Y-%m-%d %H:%M') if p.fecha_registro else 'N/A',
                       parse_monto_string(p.monto),
                       parse_monto_string(p.tasa_cambio),
                       parse_monto_string(p.monto_usd)])
            r = ws.max_row
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = thin_border
                ws.cell(row=r, column=c).font = normal_font
                if c >= 4:
                    ws.cell(row=r, column=c).alignment = right_align
                    ws.cell(row=r, column=c).number_format = '#,##0.00'
    else:
        ws.append(['Sin movimientos', '-', '-', '-', '-', '-'])
        for c in range(1, 7):
            ws.cell(row=ws.max_row, column=c).alignment = center_align

    # --- Ajuste de anchos de columna ---
    col_widths = {1: 22, 2: 28, 3: 18, 4: 16, 5: 14, 6: 14}
    for col, width in col_widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def _crear_pdf_reporte(resultados: List[dict], pagos_detalle: List[models.Pago], tipo_reporte: str, start_date: Optional[datetime], end_date: Optional[datetime], empresa_nombre: str = '', color_primario: str = '#1e3a8a', color_secundario: str = '#dbeafe', logo_bytes: Optional[bytes] = None, rif: str = '', contacto: str = '') -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.units import inch
    from PIL import Image as PILImage
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    normal_style = styles['Normal']
    section_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], spaceAfter=10, spaceBefore=15)
    meta_style = ParagraphStyle('MetaData', parent=normal_style, fontSize=9, leading=12)
    prim_hex = colors.HexColor(color_primario)
    story = []
    if logo_bytes:
        try:
            pil_img = PILImage.open(io.BytesIO(logo_bytes))
            ow, oh = pil_img.size
            max_w = 1.5 * inch
            max_h = 0.75 * inch
            ratio = min(max_w/ow, max_h/oh)
            rl_img = RLImage(io.BytesIO(logo_bytes), width=ow*ratio, height=oh*ratio)
            story.append(rl_img)
            story.append(Spacer(1, 6))
        except Exception:
            pass
    titulo = empresa_nombre + ' - ' if empresa_nombre else ''
    story.append(Paragraph(f'{titulo}Reporte de Conciliación Bancaria - {tipo_reporte.title()}', title_style))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    range_parts = []
    if start_date:
        range_parts.append(f"Desde: {start_date.strftime('%Y-%m-%d')}")
    if end_date:
        range_parts.append(f"Hasta: {end_date.strftime('%Y-%m-%d')}")
    filtro_texto = ' - '.join(range_parts) if range_parts else 'Rango: completo'
    story.append(Paragraph(filtro_texto, meta_style))
    if rif:
        story.append(Paragraph(f'RIF: {rif}', meta_style))
    if contacto:
        story.append(Paragraph(f'Contacto: {contacto}', meta_style))
    story.append(Spacer(1, 15))
    story.append(Paragraph('Resumen Agregado', section_style))
    periodo_style = ParagraphStyle('PeriodoCell', parent=normal_style, fontName='Helvetica', fontSize=8, leading=10, wordWrap='CJK')
    data = [['Periodo', 'Desde', 'Hasta', 'Total Bs', 'Total USD', 'Conteo']]
    for item in resultados:
        data.append([Paragraph(_limpiar_periodo_texto(item['periodo']), periodo_style), item['desde'].strftime('%Y-%m-%d') if item['desde'] else '', item['hasta'].strftime('%Y-%m-%d') if item['hasta'] else '', f"{parse_monto_string(item.get('total_bs')):.2f}", f"{parse_monto_string(item.get('total_usd')):.2f}", str(item.get('conteo', 0))])
    totales = _agregar_total_reporte(resultados)
    data.append(['Totales', '', '', f"{totales['total_bs']:.2f}", f"{totales['total_usd']:.2f}", str(totales['total_pagos'])])
    table = Table(data, colWidths=[120, 65, 65, 75, 75, 50])
    style = [('BACKGROUND', (0, 0), (-1, 0), prim_hex), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('ALIGN', (3, 1), (4, -1), 'RIGHT'), ('ALIGN', (5, 1), (5, -1), 'RIGHT'), ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1'))]
    for row_idx in range(1, len(data)):
        if row_idx % 2 == 1:
            style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f3f4f6')))
    table.setStyle(TableStyle(style))
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(Paragraph('Detalle Individual de Pagos', section_style))
    data_det = [['Referencia', 'Banco', 'Fecha', 'Monto Bs', 'Tasa ($)', 'Monto USD']]
    for p in pagos_detalle:
        data_det.append([p.referencia, p.banco or '-', p.fecha_registro.strftime('%Y-%m-%d %H:%M') if p.fecha_registro else 'N/A', f'{parse_monto_string(p.monto):.2f}', f'{parse_monto_string(p.tasa_cambio):.2f}', f'{parse_monto_string(p.monto_usd):.2f}'])
    if len(data_det) == 1:
        data_det.append(['Sin movimientos', '-', '-', '-', '-', '-'])
    table_det = Table(data_det, colWidths=[90, 130, 85, 75, 60, 75])
    style_det = [('BACKGROUND', (0, 0), (-1, 0), prim_hex), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('ALIGN', (3, 1), (5, -1), 'RIGHT'), ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1'))]
    for row_idx in range(1, len(data_det)):
        if row_idx % 2 == 1:
            style_det.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f3f4f6')))
    table_det.setStyle(TableStyle(style_det))
    story.append(table_det)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def _crear_nombre_archivo(tipo_reporte: str, formato: str) -> str:
    sufijo = datetime.now().strftime('%Y%m%d%H%M%S')
    return f'reportes-{tipo_reporte}-{sufijo}.{formato}'

def _parse_vision_response(data: dict) -> dict:
    if not data:
        return {}
    return {
        'banco_predicho': data.get('banco') or data.get('banco_predicho') or data.get('banco_ia'),
        'sudeban_code': data.get('sudeban_code'),
    }

async def _detectar_banco_con_vision(image_bytes: bytes) -> dict:
    from ai_client import openrouter
    if not openrouter.is_available():
        return {}
    prompt_text = 'Extrae banco_predicho y sudeban_code de este comprobante. JSON solido.'
    try:
        result = await openrouter.analyze_image(image_bytes, prompt_text)
        if result:
            return _parse_vision_response(result)
        return {}
    except Exception as e:
        logger.warning('OpenRouter Vision fallo: %s', e)
    return {}

async def _extraer_datos_vision(image_bytes: bytes) -> dict | None:
    """Envía imagen a OpenRouter Vision y extrae monto, referencia, banco, cedula y sudeban_code."""
    from ai_client import openrouter
    if not openrouter.is_available():
        return None
    prompt_text = "Extrae monto (float), referencia (digitos), banco (nombre), cedula (string), sudeban_code (4 digitos) de este comprobante. JSON solido. null si no visible."
    try:
        result = await openrouter.analyze_image(image_bytes, prompt_text)
        if result:
            logger.info('Vision extrajo: ref=%s monto=%s banco=%s', result.get('referencia'), result.get('monto'), result.get('banco'))
        else:
            logger.warning('Vision no pudo extraer JSON')
        return result
    except Exception as e:
        logger.warning('OpenRouter Vision extraccion fallo: %s', e)
    return None
