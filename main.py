import csv
import os
import re
import sys
import shutil
import uuid
import logging
import hashlib
import traceback
import io
import tempfile
import sqlite3
import base64
import json
from enum import Enum
from pathlib import Path
from typing import Optional, List

from PIL import Image
from dotenv import load_dotenv, set_key
from contextlib import asynccontextmanager

from fastapi import FastAPI, UploadFile, File, Form, Depends, HTTPException, Header, Body
from fastapi.responses import JSONResponse, FileResponse, Response, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from logging.handlers import RotatingFileHandler
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, desc, exc as sa_exc, inspect, text, or_
from sqlalchemy.engine.url import make_url
from datetime import datetime, timedelta
from pydantic import BaseModel, field_validator

from exchange import get_tasa_bcv, convertir_payments, TasaNoDisponibleError

# Carga las variables de entorno desde un archivo .env ANTES de importar módulos locales
base_dir = Path(sys.executable if getattr(sys, 'frozen', False) else __file__).resolve().parent
dotenv_path = base_dir / '.env'
load_dotenv(dotenv_path=dotenv_path)

uploads_dir = base_dir / 'uploads'
static_dir = base_dir / 'static'
uploads_dir.mkdir(parents=True, exist_ok=True)
static_dir.mkdir(parents=True, exist_ok=True)

# Importamos tus módulos (ahora database.py es más robusto)
from database import engine, get_db, Base, SQLALCHEMY_DATABASE_URL
import models
import ocr_engine  # Tu motor de OCR
import bank_rules
def _setup_logging() -> logging.Logger:
    """Configura un logger robusto que escribe en consola y en archivos rotativos."""
    os.makedirs("logs", exist_ok=True)
    logger = logging.getLogger("ocr_api")
    if logger.handlers:
        return logger

    logger.setLevel(os.getenv("LOG_LEVEL", "INFO").upper())
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(name)s - %(message)s")

    file_handler = RotatingFileHandler(
        os.path.join("logs", "app.log"),
        maxBytes=2 * 1024 * 1024, # Archivos de 2MB
        backupCount=5,
        encoding="utf-8",
    )
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)

    console = logging.StreamHandler()
    console.setFormatter(fmt)
    logger.addHandler(console)

    return logger

logger = _setup_logging()

# --- MEJORA: Uso de Lifespan para inicialización limpia ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Maneja el inicio y cierre de la aplicación de forma segura."""
    try:
        Base.metadata.create_all(bind=engine)
        inspector = inspect(engine)
        if "pagos" in inspector.get_table_names():
            columnas_pagos = [col["name"] for col in inspector.get_columns("pagos")]
            if "tasa_momento" not in columnas_pagos:
                with engine.begin() as conn:
                    conn.execute(text('ALTER TABLE pagos ADD COLUMN tasa_momento FLOAT'))
                logger.info("Se agregó la columna 'tasa_momento' en la tabla pagos.")
        logger.info("Sistema de Base de Datos inicializado: Tablas verificadas.")
        
        logger.info("🚀 [SISTEMA] Arquitectura OCR: RapidOCR (Motor Único)")
    except Exception as e:
        logger.error("❌ Error crítico al inicializar la base de datos: %s", e)
    yield

# --- MEJORA: Mover la documentación de la API para liberar la raíz ---
app = FastAPI(
    docs_url="/api/docs", 
    redoc_url=None, 
    openapi_url="/api/openapi.json",
    lifespan=lifespan
)

# --- MEJORA: Manejadores de errores de base de datos ---
@app.exception_handler(sa_exc.OperationalError)
def handle_db_operational_error(request, exc):
    logger.critical("Error de conexión con la base de datos: %s", exc)
    return JSONResponse(
        status_code=503,
        content={"detail": "Base de datos no disponible. Revisa la configuración en tu .env."},
    )

@app.exception_handler(UnicodeDecodeError)
def handle_db_unicode_decode_error(request, exc):
    logger.critical("Error de codificación con la base de datos (común en Windows): %s", exc)
    return JSONResponse(
        status_code=503,
        content={"detail": "Error de codificación con la BD. Revisa credenciales y considera definir PGCLIENTENCODING en .env."},
    )

# Montar carpetas estáticas
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")
app.mount("/uploads", StaticFiles(directory=str(uploads_dir)), name="uploads")

# --- RUTA PARA FAVICON ---
@app.get('/favicon.ico', include_in_schema=False)
async def favicon():
    favicon_path = str(static_dir / "favicon.ico")
    if os.path.exists(favicon_path):
        return FileResponse(favicon_path)
    else:
        # Si no tienes un favicon, devuelve una respuesta vacía para evitar el error 404 en los logs.
        # El navegador no volverá a pedirlo en esta sesión.
        return Response(status_code=204)

# --- MEJORA: Endpoint de Health Check ---
@app.get("/healthz", include_in_schema=False)
def healthz():
    """Verifica que el servidor está en línea."""
    return {"status": "ok"}

# --- MEJORA: Función para requerir API Key opcional ---
def require_api_key(x_api_key: Optional[str] = Header(None)):
    configured_key = os.getenv("API_KEY")
    if configured_key and (not x_api_key or x_api_key != configured_key):
        raise HTTPException(status_code=401, detail="API key inválida o no proporcionada")
    return True

# --- MODELOS PYDANTIC (Para validar datos que entran) ---

# Modelo para el estado del pago, asegura que solo se usen valores válidos
class EstadoPago(str, Enum):
    no_verificado = "no_verificado"
    verificado = "verificado"
    falso = "falso"

class EstadoUpdate(BaseModel):
    estado: EstadoPago

class VisionBankDetectionRequest(BaseModel):
    image_base64: str

# Modelos para Clientes
class ClienteBase(BaseModel):
    nombre: str
    cedula: str
    telefono: Optional[str] = None

    @field_validator('cedula', 'telefono')
    @classmethod
    def check_numeric(cls, v):
        if v is not None and v != "":
            v = v.strip()
            if not v.isdigit():
                raise ValueError('Este campo debe contener solo números')
        return v

class Cliente(ClienteBase):
    id: int
    class Config:
        from_attributes = True

# Modelo para mostrar un pago dentro de la lista de historial de un cliente
class PagoParaCliente(BaseModel):
    id: int
    referencia: str
    monto: float
    monto_usd: Optional[float] = None
    tasa_cambio: Optional[float] = None
    fecha_registro: datetime
    estado: str

    class Config:
        from_attributes = True

# Modelo para un pago con datos completos y filtros de banco
class PagoResponse(BaseModel):
    id: int
    referencia: str
    banco: str
    banco_destino: Optional[str] = None
    monto: float
    monto_usd: Optional[float] = None
    tasa_momento: Optional[float] = None
    tasa_cambio: Optional[float] = None
    fecha_registro: datetime
    estado: str
    cliente_id: Optional[int] = None
    cliente: Optional[Cliente] = None
    ruta_imagen: Optional[str] = None

    class Config:
        from_attributes = True

# Modelo de respuesta para un cliente con su historial de pagos
class ClienteConPagos(Cliente):
    pagos: List[PagoParaCliente] = []
    total_bs: float = 0.0
    total_usd: float = 0.0
    total_pagos: int = 0


class ReporteResumen(BaseModel):
    periodo: str
    desde: datetime
    hasta: datetime
    total_bs: float
    total_usd: float
    conteo: int

class ReporteResponse(BaseModel):
    tipo_reporte: str
    resultados: List[ReporteResumen]
    total_bs: float
    total_usd: float
    total_pagos: int

class PagosResponse(BaseModel):
    items: List[PagoResponse]
    total: int
    page: int
    pages: int

class PagoManual(BaseModel):
    banco: str
    referencia: str
    monto: float
    cliente_id: Optional[int] = None

    @field_validator('banco', 'referencia', mode='before')
    def validar_texto_requerido(cls, valor, info):
        if isinstance(valor, str):
            valor = valor.strip()
        if not valor:
            field_name = info.field_name.capitalize()
            raise ValueError(f"{field_name} es obligatorio y no puede estar vacío.")
        return valor

    @field_validator('referencia')
    def validar_referencia(cls, valor):
        if isinstance(valor, str) and not valor.isdigit():
            raise ValueError("La referencia debe contener solo números.")
        return valor

    @field_validator('monto')
    def validar_monto(cls, valor):
        if valor is None:
            raise ValueError("El monto es obligatorio y debe ser un número mayor a cero.")
        try:
            monto = float(valor)
        except (TypeError, ValueError):
            raise ValueError("El monto debe ser un número válido mayor a cero.")
        if monto <= 0:
            raise ValueError("El monto debe ser mayor a cero.")
        return monto

class ConversionRequest(BaseModel):
    monto_bs: float

class ConversionResponse(BaseModel):
    monto_bs: float
    tasa_bcv: float
    fecha_consulta: datetime
    monto_usd: float
    origen: str
    es_fallback: bool = False

class GestionApiKey(BaseModel):
    api_key: str

class GestionCredentials(BaseModel):
    admin_user: str
    admin_pass: str

class ConfirmBody(BaseModel):
    confirm: bool

class TasaBCVUpdate(BaseModel):
    tasa_bcv: float

class ChatQuery(BaseModel):
    pregunta: str

# --- FUNCIONES AUXILIARES ---
def registrar_auditoria(db: Session, pago_id: int, accion: str, detalles: str):
    """Función para ahorrar trabajo: registra movimientos en el historial"""
    nuevo_historial = models.PagoHistory(
        pago_id=pago_id,
        accion=accion,
        detalles=detalles,
        usuario="sistema_ia"
    )
    db.add(nuevo_historial)
    db.commit()


def _get_env_path() -> Path:
    return base_dir / '.env'


def _update_env_variable(key: str, value: str):
    env_path = _get_env_path()
    env_path.parent.mkdir(parents=True, exist_ok=True)
    set_key(str(env_path), key, value)
    os.environ[key] = value
    return True


def _get_database_type() -> str:
    url = SQLALCHEMY_DATABASE_URL.lower()
    if url.startswith('sqlite://'):
        return 'sqlite'
    if url.startswith('postgresql://') or url.startswith('postgres://'):
        return 'postgresql'
    return 'unsupported'


def _get_sqlite_db_path() -> Optional[Path]:
    if _get_database_type() != 'sqlite':
        return None
    if SQLALCHEMY_DATABASE_URL.startswith('sqlite:///'):
        sqlite_path = SQLALCHEMY_DATABASE_URL.replace('sqlite:///', '', 1)
        return Path(sqlite_path).resolve()
    if SQLALCHEMY_DATABASE_URL.startswith('sqlite://'):
        sqlite_path = SQLALCHEMY_DATABASE_URL.replace('sqlite://', '', 1)
        return Path(sqlite_path).resolve()
    return None


# --- RUTAS ---

# Redirecciones útiles para la UI del frontend
@app.get("/", include_in_schema=False)
async def root_redirect():
    """Redirige a la interfaz web estática."""
    return RedirectResponse(url="/static/index.html")


@app.get("/panel")
async def panel_redirect():
    """Endpoint visible en la documentación que abre la vista del panel web."""
    return RedirectResponse(url="/static/index.html")

# --- NUEVAS RUTAS PARA CLIENTES ---

@app.post("/clientes/", response_model=Cliente, status_code=201)
def crear_cliente(cliente: ClienteBase, db: Session = Depends(get_db)):
    """Crea un nuevo cliente en la base de datos."""
    db_cliente = db.query(models.Cliente).filter(models.Cliente.cedula == cliente.cedula).first()
    if db_cliente:
        raise HTTPException(status_code=409, detail=f"Un cliente con la cédula {cliente.cedula} ya existe.")
    
    nuevo_cliente = models.Cliente(nombre=cliente.nombre, cedula=cliente.cedula, telefono=cliente.telefono)
    db.add(nuevo_cliente)
    db.commit()
    db.refresh(nuevo_cliente)
    return nuevo_cliente

@app.get("/clientes/", response_model=List[Cliente])
def leer_clientes(q: Optional[str] = None, db: Session = Depends(get_db)):
    """Obtiene una lista de todos los clientes, opcionalmente filtrada."""
    query = db.query(models.Cliente)
    if q:
        query = query.filter(
            (models.Cliente.nombre.ilike(f"%{q}%")) | 
            (models.Cliente.cedula.ilike(f"%{q}%")) |
            (models.Cliente.telefono.ilike(f"%{q}%"))
        )
    return query.all()

@app.put("/clientes/{cliente_id}", response_model=Cliente)
def actualizar_cliente(cliente_id: int, cliente_data: ClienteBase, db: Session = Depends(get_db)):
    db_cliente = db.query(models.Cliente).filter(models.Cliente.id == cliente_id).first()
    if not db_cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    db_cliente.nombre = cliente_data.nombre
    db_cliente.cedula = cliente_data.cedula
    db_cliente.telefono = cliente_data.telefono
    
    db.commit()
    db.refresh(db_cliente)
    return db_cliente

@app.put("/clientes/{cliente_id}/", include_in_schema=False)
def actualizar_cliente_trailing_slash(cliente_id: int, cliente_data: ClienteBase, db: Session = Depends(get_db)):
    return actualizar_cliente(cliente_id, cliente_data, db)

@app.delete("/clientes/{cliente_id}")
def eliminar_cliente(cliente_id: int, db: Session = Depends(get_db)):
    db_cliente = db.query(models.Cliente).filter(models.Cliente.id == cliente_id).first()
    if not db_cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    db.delete(db_cliente)
    db.commit()
    return {"mensaje": "Cliente eliminado"}

@app.delete("/clientes/{cliente_id}/", include_in_schema=False)
def eliminar_cliente_trailing_slash(cliente_id: int, db: Session = Depends(get_db)):
    return eliminar_cliente(cliente_id, db)

@app.get("/clientes/{cliente_id}/pagos", response_model=ClienteConPagos)
def leer_pagos_de_cliente(cliente_id: int, db: Session = Depends(get_db)):
    """Obtiene los detalles de un cliente y todos sus pagos asociados."""
    cliente = db.query(models.Cliente).filter(models.Cliente.id == cliente_id).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    pagos = cliente.pagos or []
    total_bs = sum(float(p.monto or 0.0) for p in pagos)
    total_usd = sum(float(p.monto_usd or 0.0) for p in pagos)
    total_pagos = len(pagos)

    return {
        "id": cliente.id,
        "nombre": cliente.nombre,
        "cedula": cliente.cedula,
        "telefono": cliente.telefono,
        "pagos": pagos,
        "total_bs": total_bs,
        "total_usd": total_usd,
        "total_pagos": total_pagos,
    }

@app.post("/IA/consultar/")
async def consultar_datos_ia(query: ChatQuery, db: Session = Depends(get_db)):
    """Permite hacer preguntas sobre los pagos en lenguaje natural."""
    # 1. Consultas rápidas para contexto de negocio
    total_clientes = db.query(models.Cliente).count()
    total_pagos = db.query(models.Pago).count()
    total_historial = db.query(models.PagoHistory).count()

    # Distribución de pagos por estado
    pagos_por_estado = db.query(
        models.Pago.estado,
        func.count(models.Pago.id).label('conteo'),
        func.sum(models.Pago.monto).label('total_monto')
    ).group_by(models.Pago.estado).all()
    contexto_por_estado = "\n".join([
        f"- {estado}: {conteo} pagos, total {total_monto or 0.0:,.2f} Bs"
        for estado, conteo, total_monto in pagos_por_estado
    ]) or "- Sin registros de pago por estado."

    # Mejor cliente (más pagos realizados)
    mejor_cliente = db.query(
        models.Cliente.nombre,
        func.count(models.Pago.id).label('total')
    ).join(models.Pago).group_by(models.Cliente.id).order_by(text('total DESC')).first()
    nombre_mejor = mejor_cliente.nombre if mejor_cliente else "N/A"

    # Recaudación reciente
    inicio_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0)
    recaudado_mes = db.query(func.sum(models.Pago.monto)).filter(models.Pago.fecha_registro >= inicio_mes).scalar() or 0.0
    inicio_30_dias = datetime.now() - timedelta(days=30)
    recaudado_30_dias = db.query(func.sum(models.Pago.monto)).filter(models.Pago.fecha_registro >= inicio_30_dias).scalar() or 0.0
    inicio_60_dias = datetime.now() - timedelta(days=60)
    recaudado_60_dias = db.query(func.sum(models.Pago.monto)).filter(models.Pago.fecha_registro >= inicio_60_dias).scalar() or 0.0

    # Últimos 5 pagos
    ultimos_pagos = db.query(models.Pago).order_by(desc(models.Pago.id)).limit(5).all()
    contexto_pagos = "\n".join([
        f"- Ref: {p.referencia}, Banco: {p.banco}, Monto: {p.monto} Bs, Cliente: {p.cliente.nombre if p.cliente else 'Particular'}"
        for p in ultimos_pagos
    ]) or "- No hay pagos recientes registrados."

    # Últimas 5 acciones de historial
    ultimas_acciones = db.query(models.PagoHistory).order_by(desc(models.PagoHistory.id)).limit(5).all()
    contexto_historial = "\n".join([
        f"- Pago ID: {h.pago_id}, Acción: {h.accion}, Usuario: {h.usuario or 'desconocido'}, Detalles: {h.detalles}"
        for h in ultimas_acciones
    ]) or "- No hay acciones de historial recientes."

    # Última tasa BCV registrada
    ultima_tasa = db.query(models.TasaCambio).order_by(desc(models.TasaCambio.fecha_actualizacion)).first()
    tasa_bcv_texto = f"{ultima_tasa.monto_tasa:.4f} (actualizada {ultima_tasa.fecha_actualizacion})" if ultima_tasa else "No disponible"

    total_usd_recaudado = db.query(func.sum(models.Pago.monto_usd)).scalar() or 0.0
    tasa_promedio_pagos = db.query(func.avg(models.Pago.tasa_momento)).scalar() or 0.0
    balance_total_bs = db.query(func.sum(models.Pago.monto)).scalar() or 0.0

    prompt = f"""
    Eres un asistente contable experto del Sistema de Conciliación.
    TABLAS Y ESQUEMA DISPONIBLES EN LA BASE DE DATOS:
    - clientes(id, nombre, cedula, telefono)
    - pagos(id, referencia, banco, banco_destino, monto, monto_usd, tasa_momento, tasa_cambio, fecha_registro, ruta_imagen, file_hash, estado, cliente_id)
    - pagos_history(id, pago_id, accion, detalles, usuario, fecha)
    - tasas_cambio(id, proveedor, monto_tasa, fecha_actualizacion)
    RELACIONES:
    - pago.cliente_id -> cliente.id
    - un pago puede no tener cliente asociado

    CONTEXTO AGREGADO:
    - Clientes totales: {total_clientes}
    - Pagos totales: {total_pagos}
    - Registros de auditoría: {total_historial}
    - Pagos por estado:\n{contexto_por_estado}
    - Total recaudado este mes: {recaudado_mes:,.2f} Bs.
    - Total recaudado último 30 días: {recaudado_30_dias:,.2f} Bs.
    - Total recaudado últimos 60 días: {recaudado_60_dias:,.2f} Bs.
    - Total recaudado en USD (sumatoria de monto_usd): {total_usd_recaudado:,.2f} USD.
    - Tasa promedio de los últimos pagos: {tasa_promedio_pagos:.4f}.
    - Balance total en Bs: {balance_total_bs:,.2f} Bs.
    - Última tasa BCV registrada: {tasa_bcv_texto}
    - Mejor cliente (frecuencia): {nombre_mejor}

    ÚLTIMOS 5 PAGOS REGISTRADOS:
    {contexto_pagos}

    ÚLTIMAS 5 ACCIONES DE HISTORIAL:
    {contexto_historial}

    INSTRUCCIONES:
    - Usa SOLO los datos proporcionados en este prompt.
    - No inventes valores, no supongas pagos, clientes ni datos adicionales.
    - Si no tienes información suficiente para responder, di "No hay suficiente información en los datos".
    - Responde de forma breve y profesional.

    PREGUNTA DEL USUARIO:
    {query.pregunta}
    """

    try:
        from groq import Groq
        client = Groq(api_key=os.getenv("GROQ_API_KEY"))
        response = client.chat.completions.create(
            messages=[{"role": "user", "content": prompt}],
            model=os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile"),
            temperature=0.2
        )
        return {"respuesta": response.choices[0].message.content}
    except Exception as e:
        return {"respuesta": "No fue posible procesar la consulta IA en este momento. Comprueba la configuración de GROQ_API_KEY y vuelve a intentarlo."}

@app.post("/convertir-a-usd/", response_model=ConversionResponse)
async def convertir_monto_a_usd(data: ConversionRequest, db: Session = Depends(get_db)):
    """Convierte un monto en bolívares a USD usando la tasa BCV con fallback robusto."""
    try:
        result = await convertir_payments(db, data.monto_bs)
    except TasaNoDisponibleError as e:
        raise HTTPException(status_code=503, detail=str(e))

    return {
        "monto_bs": float(result["monto_bs"]),
        "tasa_bcv": float(result["tasa_bcv"]),
        "fecha_consulta": result["fecha"],
        "monto_usd": float(result["monto_usd"]),
        "origen": result["origen"],
        "es_fallback": result["origen"] != "API",
    }

@app.get("/tasa-bcv/")
async def obtener_tasa_bcv_endpoint(db: Session = Depends(get_db)):
    tasa_info = await get_tasa_bcv(db)
    tasa = float(tasa_info["tasa"])
    origen = tasa_info["origen"]
    fecha = tasa_info["fecha"]
    es_fallback = origen != "API"
    return {"tasa_bcv": tasa, "fecha_consulta": fecha, "origen": origen, "es_fallback": es_fallback}

@app.get("/bancos/")
def listar_bancos():
    return {"bancos": bank_rules.get_available_banks()}

def _verificar_estado_groq(api_key: str) -> tuple[bool, str]:
    if not api_key:
        return False, "No se ha configurado la clave Groq."

    try:
        from groq import Groq
        client = Groq(api_key=api_key, timeout=3.0)
        # Realiza una llamada mínima al servicio para verificar conectividad.
        client.chat.completions.create(
            messages=[{"role": "user", "content": "Responde con pong"}],
            model=os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile"),
            temperature=0.0,
            max_tokens=1,
            timeout=3.0,
        )
        return True, "Clave Groq cargada y verificada."
    except Exception as e:
        logger.warning("No se pudo verificar Groq API: %s", e)
        return False, "No se puede conectar a Groq. Comprueba tu conexión a internet y la clave de API."

@app.get("/gestion/ia/status")
def estado_groq_api():
    api_key = os.getenv("GROQ_API_KEY", "").strip()
    online, message = _verificar_estado_groq(api_key)
    state = "online" if online else ("invalid_key" if not api_key else "offline")
    return {
        "state": state,
        "api_key": api_key,
        "message": message
    }

@app.post("/gestion/ia/key")
def guardar_groq_api_key(data: GestionApiKey):
    api_key = data.api_key.strip()
    if not api_key:
        raise HTTPException(status_code=400, detail="La clave de Groq no puede estar vacía.")

    _update_env_variable("GROQ_API_KEY", api_key)
    online, message = _verificar_estado_groq(api_key)
    state = "online" if online else "offline"
    return {"mensaje": "Clave Groq guardada correctamente.", "state": state, "message": message}

@app.get("/gestion/db/status")
def estado_db():
    db_type = _get_database_type()
    if db_type == 'sqlite':
        db_path = _get_sqlite_db_path()
        return {
            "database_type": "sqlite",
            "path": str(db_path) if db_path else None,
            "exists": bool(db_path and db_path.exists()),
            "message": "SQLite local detectada. Importación de .db disponible."
        }
    if db_type == 'postgresql':
        return {
            "database_type": "postgresql",
            "message": "PostgreSQL detectada. No hay archivo SQLite local. Solo la importación de pagos CSV está disponible.",
            "info": "PostgreSQL"
        }
    return {"database_type": "unsupported", "message": "Tipo de base de datos no soportado para gestión automática.", "info": "No disponible"}

@app.get("/gestion/db/export-pagos")
def exportar_pagos_csv(db: Session = Depends(get_db)):
    pagos = db.query(models.Pago).order_by(models.Pago.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "referencia", "banco", "banco_destino", "monto", "monto_usd", "tasa_momento", "tasa_cambio", "fecha_registro", "estado", "cliente_id", "cliente_nombre", "ruta_imagen"])
    for pago in pagos:
        writer.writerow([
            pago.id,
            pago.referencia,
            pago.banco,
            pago.banco_destino or "",
            pago.monto,
            pago.monto_usd,
            pago.tasa_momento,
            pago.tasa_cambio,
            pago.fecha_registro.isoformat() if pago.fecha_registro else "",
            pago.estado,
            pago.cliente_id,
            pago.cliente.nombre if pago.cliente else "",
            pago.ruta_imagen or ""
        ])
    return Response(output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=pagos_export.csv"})

@app.post("/gestion/db/import-pagos")
def importar_pagos_csv(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith('.csv'):
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV.")

    contenido = file.file.read().decode('utf-8', errors='replace')
    lector = csv.DictReader(io.StringIO(contenido))
    creados = 0
    actualizados = 0
    omitidos = 0
    for fila in lector:
        referencia = (fila.get('referencia') or fila.get('Referencia') or '').strip()
        banco = (fila.get('banco') or fila.get('Banco') or '').strip()
        monto_texto = (fila.get('monto') or fila.get('Monto') or '').strip()
        if not referencia or not banco or not monto_texto:
            omitidos += 1
            continue
        try:
            monto = float(monto_texto.replace(',', '.'))
        except ValueError:
            omitidos += 1
            continue

        pago_existente = db.query(models.Pago).filter(
            models.Pago.referencia == referencia,
            models.Pago.banco == banco
        ).first()
        if pago_existente:
            omitidos += 1
            continue

        cliente = None
        cedula = (fila.get('cliente_cedula') or fila.get('cedula') or fila.get('Cedula') or '').strip()
        nombre_cliente = (fila.get('cliente_nombre') or fila.get('cliente') or fila.get('Cliente') or '').strip()
        telefono = (fila.get('telefono') or fila.get('Telefono') or fila.get('telefono_cliente') or '').strip()
        if cedula:
            cliente = db.query(models.Cliente).filter(models.Cliente.cedula == cedula).first()
            if not cliente and nombre_cliente:
                cliente = models.Cliente(nombre=nombre_cliente, cedula=cedula, telefono=telefono or None)
                db.add(cliente)
                db.flush()

        fecha_registro = None
        fecha_texto = (fila.get('fecha_registro') or fila.get('fecha') or fila.get('Fecha') or '').strip()
        if fecha_texto:
            for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                try:
                    fecha_registro = datetime.strptime(fecha_texto, fmt)
                    break
                except ValueError:
                    continue

        monto_usd_texto = (fila.get('monto_usd') or fila.get('Monto_USD') or fila.get('montoUsd') or '').strip()
        tasa_momento_texto = (fila.get('tasa_momento') or fila.get('Tasa_Momento') or '').strip()
        tasa_cambio_texto = (fila.get('tasa_cambio') or fila.get('Tasa_Cambio') or '').strip()
        estado = (fila.get('estado') or fila.get('Estado') or 'no_verificado').strip() or 'no_verificado'
        banco_destino = (fila.get('banco_destino') or fila.get('Banco_Destino') or '').strip() or None
        ruta_imagen = (fila.get('ruta_imagen') or fila.get('Ruta_Imagen') or '').strip() or None

        def parse_float(valor, fallback=None):
            if not valor:
                return fallback
            try:
                return float(valor.replace(',', '.'))
            except ValueError:
                return fallback

        pago = models.Pago(
            referencia=referencia,
            banco=banco,
            banco_destino=banco_destino,
            monto=monto,
            monto_usd=parse_float(monto_usd_texto, 0.0),
            tasa_momento=parse_float(tasa_momento_texto, 1.0),
            tasa_cambio=parse_float(tasa_cambio_texto, 1.0),
            fecha_registro=fecha_registro,
            ruta_imagen=ruta_imagen,
            estado=estado,
            cliente_id=cliente.id if cliente else None,
        )
        db.add(pago)
        creados += 1

    db.commit()
    return {
        "mensaje": f"Pagos importados: {creados}, omitidos: {omitidos}.",
        "creados": creados,
        "omitidos": omitidos
    }

@app.post("/gestion/db/clear-test-data")
def limpiar_datos_prueba(data: ConfirmBody, db: Session = Depends(get_db)):
    if not data.confirm:
        raise HTTPException(status_code=400, detail="Se requiere confirmación para borrar datos de prueba.")
    pagos = db.query(models.Pago).filter(
        or_(
            models.Pago.estado == 'no_verificado',
            models.Pago.referencia.ilike('%test%')
        )
    ).all()
    deleted = len(pagos)
    for pago in pagos:
        db.delete(pago)
    db.commit()
    return {"mensaje": f"{deleted} pagos de prueba eliminados correctamente."}

@app.post("/gestion/clientes/clear")
def limpiar_clientes(data: ConfirmBody, db: Session = Depends(get_db)):
    if not data.confirm:
        raise HTTPException(status_code=400, detail="Se requiere confirmación para borrar los clientes.")
    clientes = db.query(models.Cliente).all()
    deleted = len(clientes)
    for cliente in clientes:
        db.delete(cliente)
    db.commit()
    return {"mensaje": f"{deleted} clientes eliminados correctamente."}

@app.get("/gestion/db/credentials")
def obtener_credenciales():
    return {
        "admin_user": os.getenv("ADMIN_USER", ""),
        "admin_pass": os.getenv("ADMIN_PASS", "")
    }

@app.post("/gestion/db/credentials")
def guardar_credenciales(data: GestionCredentials):
    _update_env_variable("ADMIN_USER", data.admin_user.strip())
    _update_env_variable("ADMIN_PASS", data.admin_pass.strip())
    return {"mensaje": "Credenciales guardadas correctamente."}

@app.get("/gestion/clientes/export")
def exportar_clientes_csv(db: Session = Depends(get_db)):
    clientes = db.query(models.Cliente).order_by(models.Cliente.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["nombre", "cedula", "telefono"])
    for cliente in clientes:
        writer.writerow([cliente.nombre, cliente.cedula, cliente.telefono or ""])
    return Response(output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=clientes_export.csv"})

@app.post("/gestion/clientes/import")
def importar_clientes(archivo: UploadFile = File(...), db: Session = Depends(get_db)):
    if not archivo.filename.lower().endswith('.csv'):
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV.")
    contenido = archivo.file.read().decode('utf-8', errors='replace')
    lector = csv.DictReader(io.StringIO(contenido))
    creados = 0
    actualizados = 0
    for fila in lector:
        nombre = (fila.get('nombre') or fila.get('Nombre') or '').strip()
        cedula = (fila.get('cedula') or fila.get('Cédula') or fila.get('cedula') or '').strip()
        telefono = (fila.get('telefono') or fila.get('Telefono') or fila.get('teléfono') or '').strip()
        if not nombre or not cedula:
            continue
        existente = db.query(models.Cliente).filter(models.Cliente.cedula == cedula).first()
        if existente:
            existente.nombre = nombre
            existente.telefono = telefono
            actualizados += 1
        else:
            db.add(models.Cliente(nombre=nombre, cedula=cedula, telefono=telefono))
            creados += 1
    db.commit()
    return {"mensaje": f"Clientes importados: {creados}, actualizados: {actualizados}", "creados": creados, "actualizados": actualizados}

@app.get("/gestion/clientes/summary")
def resumen_clientes(db: Session = Depends(get_db)):
    total = db.query(models.Cliente).count()
    ultimos = db.query(models.Cliente).order_by(desc(models.Cliente.id)).limit(5).all()
    return {
        "total": total,
        "ultimos": [{"nombre": c.nombre, "cedula": c.cedula, "telefono": c.telefono or ""} for c in ultimos]
    }

@app.post("/tasa-bcv/")
def set_tasa_bcv(data: TasaBCVUpdate, db: Session = Depends(get_db), x_api_key: Optional[str] = Header(None)):
    require_api_key(x_api_key)
    tasa_bcv = data.tasa_bcv
    if tasa_bcv <= 0:
        raise HTTPException(status_code=400, detail="La tasa debe ser mayor que cero")

    tasa_db = db.query(models.TasaCambio).filter(models.TasaCambio.proveedor == "BCV").first()
    if not tasa_db:
        tasa_db = models.TasaCambio(proveedor="BCV", monto_tasa=tasa_bcv)
        db.add(tasa_db)
    else:
        tasa_db.monto_tasa = tasa_bcv
        tasa_db.fecha_actualizacion = func.now()

    db.commit()
    return {"tasa_bcv": tasa_bcv, "mensaje": "Tasa BCV actualizada"}

@app.get("/ver-pagos/")
def leer_pagos(q: Optional[str] = None, banco: Optional[str] = None, page: int = 1, size: int = 10, limit: Optional[int] = None, offset: Optional[int] = None, db: Session = Depends(get_db)):
    # Compatibilidad: Si el frontend envía limit/offset (cache antiguo), calculamos la página
    if limit:
        size = limit
    if offset is not None:
        page = (offset // size) + 1

    # Lógica de paginación de SQL
    # Si page=1, skip=0. Si page=2, skip=10.
    skip = (page - 1) * size

    query = db.query(models.Pago)
    if q:
        query = query.filter(models.Pago.referencia.contains(q))
    if banco:
        termino = f"%{banco}%"
        query = query.filter(
            models.Pago.banco.ilike(termino)
        )

    pagos = query.options(joinedload(models.Pago.cliente)).order_by(desc(models.Pago.id)).offset(skip).limit(size).all()

    # Evitar COUNT(*) caro en tablas muy grandes; si la tabla es pequeña o la solicitud lo pide, calcularlo
    total = None
    if os.getenv("FORCE_EXACT_COUNT", "false").lower() == "true":
        total = db.query(models.Pago).count()
    elif len(pagos) < size or page == 1:
        # Para la primera página (y pocos resultados) podemos usar un COUNT real sin mucho coste
        total = db.query(models.Pago).count()
    else:
        # Aprox. by using last row index o estimación de Postgres
        try:
            total = db.execute("SELECT reltuples::BIGINT AS estimate FROM pg_class WHERE oid = 'pagos'::regclass").scalar()
            if total is None or total <= 0:
                total = db.query(models.Pago).count()
        except Exception:
            total = db.query(models.Pago).count()

    return {
        "items": pagos,
        "total": total,
        "page": page,
        "pages": (total + size - 1) // size if total is not None and total > 0 else 1
    }

@app.get("/buscar-pagos/")
def buscar_pagos(q: str, page: int = 1, size: int = 10, limit: Optional[int] = None, offset: Optional[int] = None, db: Session = Depends(get_db)):
    if limit:
        size = limit
    if offset is not None:
        page = (offset // size) + 1

    skip = (page - 1) * size
    query = db.query(models.Pago).filter(models.Pago.referencia.contains(q))

    pagos = query.options(joinedload(models.Pago.cliente)).order_by(desc(models.Pago.id)).offset(skip).limit(size).all()

    total = None
    if os.getenv("FORCE_EXACT_COUNT", "false").lower() == "true":
        total = query.count()
    elif len(pagos) < size or page == 1:
        total = query.count()
    else:
        try:
            total = db.execute("SELECT reltuples::BIGINT AS estimate FROM pg_class WHERE oid = 'pagos'::regclass").scalar()
            if total is None or total <= 0:
                total = query.count()
        except Exception:
            total = query.count()

    return {
        "items": pagos,
        "total": total,
        "page": page,
        "pages": (total + size - 1) // size if total is not None and total > 0 else 1
    }

@app.get("/pagos/", response_model=PagosResponse)
def listar_pagos(q: Optional[str] = None, banco: Optional[str] = None, page: int = 1, size: int = 10, db: Session = Depends(get_db)):
    """Lista los pagos con opción de filtrar por referencia y banco emisor u origen."""
    query = db.query(models.Pago)
    if q:
        termino_q = f"%{q}%"
        query = query.filter(models.Pago.referencia.ilike(termino_q))
    if banco:
        termino = f"%{banco}%"
        query = query.filter(
            models.Pago.banco.ilike(termino)
        )

    if page < 1:
        page = 1

    skip = (page - 1) * size
    pagos = query.options(joinedload(models.Pago.cliente)).order_by(desc(models.Pago.id)).offset(skip).limit(size).all()

    total = None
    if os.getenv("FORCE_EXACT_COUNT", "false").lower() == "true":
        total = query.count()
    elif len(pagos) < size or page == 1:
        total = query.count()
    else:
        try:
            total = db.execute("SELECT reltuples::BIGINT AS estimate FROM pg_class WHERE oid = 'pagos'::regclass").scalar()
            if total is None or total <= 0:
                total = query.count()
        except Exception:
            total = query.count()

    return {
        "items": pagos,
        "total": total,
        "page": page,
        "pages": (total + size - 1) // size if total is not None and total > 0 else 1
    }


def _agregar_total_reporte(resultado: List[dict]) -> dict:
    total_bs = sum(item["total_bs"] for item in resultado)
    total_usd = sum(item["total_usd"] for item in resultado)
    total_pagos = sum(item["conteo"] for item in resultado)
    return {
        "total_bs": total_bs,
        "total_usd": total_usd,
        "total_pagos": total_pagos,
    }


def _query_reporte(db: Session, tipo_reporte: str, fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime]) -> List[dict]:
    interval_map = {
        "diario": "day",
        "semanal": "week",
        "mensual": "month",
        "trimestral": "quarter",
        "anual": "year",
    }

    if tipo_reporte not in interval_map and tipo_reporte not in ["quincenal", "semestral"]:
        raise HTTPException(status_code=400, detail=f"Tipo de reporte desconocido: {tipo_reporte}")

    if tipo_reporte == "quincenal":
        periodo_expr = (
            func.date_trunc("month", models.Pago.fecha_registro)
            + (func.floor((func.extract("day", models.Pago.fecha_registro) - 1) / 15) * text("interval '15 days'"))
        )
    elif tipo_reporte == "semestral":
        periodo_expr = (
            func.date_trunc("year", models.Pago.fecha_registro)
            + (func.floor((func.extract("month", models.Pago.fecha_registro) - 1) / 6) * text("interval '6 months'"))
        )
    else:
        periodo_expr = func.date_trunc(interval_map[tipo_reporte], models.Pago.fecha_registro)

    query = db.query(
        periodo_expr.label("periodo"),
        func.sum(models.Pago.monto).label("total_bs"),
        func.sum(models.Pago.monto_usd).label("total_usd"),
        func.count(models.Pago.id).label("conteo"),
        func.min(models.Pago.fecha_registro).label("desde"),
        func.max(models.Pago.fecha_registro).label("hasta")
    )
    if fecha_inicio:
        query = query.filter(models.Pago.fecha_registro >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Pago.fecha_registro <= fecha_fin)

    grouped = query.group_by(periodo_expr).order_by(periodo_expr).all()

    resultados = []
    for r in grouped:
        periodo_text = _simplificar_periodo(r.periodo)
        resultados.append({
            "periodo": periodo_text,
            "desde": r.desde,
            "hasta": r.hasta,
            "total_bs": float(r.total_bs or 0),
            "total_usd": float(r.total_usd or 0),
            "conteo": int(r.conteo or 0)
        })
    return resultados


def _simplificar_periodo(periodo: Optional[object]) -> str:
    if periodo is None:
        return ""
    if isinstance(periodo, datetime):
        return periodo.strftime("%d-%m-%Y")
    if isinstance(periodo, str):
        periodo = periodo.strip()
        if not periodo:
            return ""
        try:
            if "T" in periodo:
                fecha = datetime.fromisoformat(periodo)
                return fecha.strftime("%d-%m-%Y")
        except Exception:
            pass
        if "T" in periodo:
            return periodo.split("T")[0]
        if " " in periodo and ":" in periodo:
            return periodo.split(" ")[0]
        return periodo
    return str(periodo)


def _limpiar_periodo_texto(periodo: Optional[object]) -> str:
    if periodo is None:
        return ""
    if isinstance(periodo, datetime):
        return periodo.strftime("%d-%m-%Y")
    if isinstance(periodo, str):
        texto = periodo.strip()
        if not texto:
            return ""
        match = re.search(r"(\d{2})[-/](\d{2})[-/](\d{4})", texto)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        match = re.search(r"(\d{4})-(\d{2})-(\d{2})", texto)
        if match:
            return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        texto = texto.splitlines()[0].strip()
        return texto if len(texto) <= 32 else texto[:32] + "..."
    return str(periodo)


def _query_pagos_detalle(db: Session, fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime]) -> List[models.Pago]:
    """Obtiene la lista detallada de pagos para el reporte."""
    query = db.query(models.Pago)
    if fecha_inicio:
        query = query.filter(models.Pago.fecha_registro >= fecha_inicio)
    if fecha_fin:
        query = query.filter(models.Pago.fecha_registro <= fecha_fin)
    return query.order_by(models.Pago.fecha_registro.desc()).all()


def parse_monto_string(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return 0.0

    texto = re.sub(r'\s+', '', texto)
    texto = texto.replace('Bs', '').replace('Bs.', '').replace('Bs,', '')
    texto = re.sub(r'[^0-9\,\.\-]', '', texto)

    if texto.count(',') > 0 and texto.count('.') > 0:
        if texto.rfind(',') > texto.rfind('.'):
            texto = texto.replace('.', '')
            texto = texto.replace(',', '.')
        else:
            texto = texto.replace(',', '')
    else:
        texto = texto.replace(',', '.')

    try:
        return float(texto)
    except ValueError:
        return 0.0


def _extraer_codigo_sudeban(texto: Optional[str]) -> Optional[str]:
    if not texto:
        return None
    texto = str(texto)
    match = re.search(r'\b(\d{4})\b', texto)
    if match:
        return match.group(1)
    return None


def _agrupar_totales_sudeban(pagos_detalle: List[models.Pago]) -> List[dict]:
    grupos = {}
    for pago in pagos_detalle:
        codigo = _extraer_codigo_sudeban(pago.banco) or _extraer_codigo_sudeban(pago.banco_destino)
        if codigo:
            banco_formal = bank_rules.get_bank_by_sudeban_code(codigo)
            etiqueta = f"{codigo} - {banco_formal}" if banco_formal != 'Desconocido' else f"{codigo} - Desconocido"
        else:
            etiqueta = f"Desconocido"
            codigo = "N/A"

        clave = (codigo, etiqueta)
        if clave not in grupos:
            grupos[clave] = {
                "sudeban_code": codigo,
                "banco_label": etiqueta,
                "total_bs": 0.0,
                "total_usd": 0.0,
                "conteo": 0,
            }
        grupos[clave]["total_bs"] += parse_monto_string(pago.monto)
        grupos[clave]["total_usd"] += parse_monto_string(pago.monto_usd)
        grupos[clave]["conteo"] += 1

    return [
        {
            "sudeban_code": codigo,
            "banco_label": etiqueta,
            "total_bs": valores["total_bs"],
            "total_usd": valores["total_usd"],
            "conteo": valores["conteo"],
        }
        for (codigo, etiqueta), valores in sorted(grupos.items(), key=lambda item: item[0])
    ]


def _crear_excel_reporte(resultados: List[dict], pagos_detalle: List[models.Pago], tipo_reporte: str, start_date: Optional[datetime], end_date: Optional[datetime]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    right_align = Alignment(horizontal='right')

    ws.append(["Reporte de Conciliación", tipo_reporte.title()])
    ws.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Periodo", start_date.strftime("%Y-%m-%d") if start_date else "Completo", end_date.strftime("%Y-%m-%d") if end_date else "Completo"])
    ws.append([])

    sudeban_summary = _agrupar_totales_sudeban(pagos_detalle)
    ws.append(["Código SUDEBAN", "Banco Origen", "Total Bs", "Total USD", "Conteo"])
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font

    for row in sudeban_summary:
        ws.append([
            row["sudeban_code"],
            row["banco_label"],
            row["total_bs"],
            row["total_usd"],
            row["conteo"],
        ])

    ws.append([])
    ws.append(["Resumen Agregado"])
    ws.append(["Periodo", "Desde", "Hasta", "Total Bs", "Total USD", "Conteo"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    for item in resultados:
        periodo_text = _limpiar_periodo_texto(item["periodo"])
        ws.append([
            periodo_text,
            item["desde"].strftime("%Y-%m-%d") if item["desde"] else "",
            item["hasta"].strftime("%Y-%m-%d") if item["hasta"] else "",
            parse_monto_string(item.get("total_bs")),
            parse_monto_string(item.get("total_usd")),
            item.get("conteo", 0),
        ])
        periodo_cell = ws[f"A{ws.max_row}"]
        periodo_cell.alignment = Alignment(wrap_text=True, vertical='top')

    totales = _agregar_total_reporte(resultados)
    ws.append(["Totales", "", "", totales["total_bs"], totales["total_usd"], totales["total_pagos"]])

    for sheet in wb.worksheets:
        for columna in sheet.columns:
            max_length = 0
            column_letter = columna[0].column_letter
            for cell in columna:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
            if column_letter == 'A':
                for cell in columna:
                    if cell.value is not None and isinstance(cell.value, str) and len(cell.value) > 30:
                        lines = (len(cell.value) - 1) // 30 + 1
                        current_height = sheet.row_dimensions[cell.row].height or 15
                        sheet.row_dimensions[cell.row].height = max(current_height, lines * 15)

        # Aplicar formato numérico a columnas de monto en cada hoja
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.column_letter in ('C', 'D') and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00 "Bs"'
                if cell.column_letter == 'E' and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    ws_det = wb.create_sheet(title="Detalle de Pagos")
    ws_det.append(["Referencia", "Banco Origen", "Fecha", "Monto (Bs)", "Tasa ($)", "Monto ($)"])
    for cell in ws_det[1]:
        cell.fill = header_fill
        cell.font = header_font

    for p in pagos_detalle:
        ws_det.append([
            p.referencia,
            p.banco,
            p.fecha_registro.strftime("%Y-%m-%d %H:%M") if p.fecha_registro else "N/A",
            parse_monto_string(p.monto),
            parse_monto_string(p.tasa_cambio),
            parse_monto_string(p.monto_usd),
        ])

    for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row):
        for cell in row:
            if cell.column_letter == 'D' and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "Bs"'
                cell.alignment = right_align
            if cell.column_letter in ('E', 'F') and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = right_align

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _crear_pdf_reporte(resultados: List[dict], pagos_detalle: List[models.Pago], tipo_reporte: str, start_date: Optional[datetime], end_date: Optional[datetime]) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]
    section_style = ParagraphStyle('SectionHeading', parent=styles['Heading2'], spaceAfter=10, spaceBefore=15)

    story = []
    story.append(Paragraph(f"Reporte de Conciliación Bancaria - {tipo_reporte.title()}", title_style))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))

    range_parts = []
    if start_date:
        range_parts.append(f"Desde: {start_date.strftime('%Y-%m-%d')}")
    if end_date:
        range_parts.append(f"Hasta: {end_date.strftime('%Y-%m-%d')}")
    filtro_texto = " - ".join(range_parts) if range_parts else "Rango: completo"
    story.append(Paragraph(filtro_texto, normal_style))
    story.append(Spacer(1, 15))

    sudeban_summary = _agrupar_totales_sudeban(pagos_detalle)
    story.append(Paragraph("Resumen por Código SUDEBAN", section_style))
    data_sudeban = [["Código SUDEBAN", "Banco Origen", "Total Bs", "Total USD", "Conteo"]]
    for item in sudeban_summary:
        data_sudeban.append([
            item["sudeban_code"],
            item["banco_label"],
            f"{item['total_bs']:.2f}",
            f"{item['total_usd']:.2f}",
            str(item["conteo"]),
        ])
    if len(data_sudeban) == 1:
        data_sudeban.append(["No hay datos", "", "", "", ""])

    table_sudeban = Table(data_sudeban, colWidths=[90, 160, 90, 90, 60])
    style_sudeban = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ]
    for row_idx in range(1, len(data_sudeban)):
        if row_idx % 2 == 1:
            style_sudeban.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor('#f3f4f6')))
    table_sudeban.setStyle(TableStyle(style_sudeban))
    story.append(table_sudeban)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Resumen Agregado", section_style))
    periodo_style = ParagraphStyle(
        'PeriodoCell',
        parent=normal_style,
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        wordWrap='CJK',
    )
    data = [["Periodo", "Desde", "Hasta", "Total Bs", "Total USD", "Conteo"]]
    for item in resultados:
        data.append([
            Paragraph(_limpiar_periodo_texto(item["periodo"]), periodo_style),
            item["desde"].strftime("%Y-%m-%d") if item["desde"] else "",
            item["hasta"].strftime("%Y-%m-%d") if item["hasta"] else "",
            f"{parse_monto_string(item.get('total_bs')):.2f}",
            f"{parse_monto_string(item.get('total_usd')):.2f}",
            str(item.get("conteo", 0)),
        ])
    totales = _agregar_total_reporte(resultados)
    data.append(["Totales", "", "", f"{totales['total_bs']:.2f}", f"{totales['total_usd']:.2f}", str(totales['total_pagos'])])

    table = Table(data, colWidths=[120, 65, 65, 75, 75, 50])
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (4, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ]
    for row_idx in range(1, len(data)):
        if row_idx % 2 == 1:
            style.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor('#f3f4f6')))
    table.setStyle(TableStyle(style))
    story.append(table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Detalle Individual de Pagos", section_style))
    data_det = [["Referencia", "Banco", "Fecha", "Monto Bs", "Tasa ($)", "Monto USD"]]
    for p in pagos_detalle:
        data_det.append([
            p.referencia,
            p.banco or "-",
            p.fecha_registro.strftime("%Y-%m-%d") if p.fecha_registro else "N/A",
            f"{parse_monto_string(p.monto):.2f}",
            f"{parse_monto_string(p.tasa_cambio):.2f}",
            f"{parse_monto_string(p.monto_usd):.2f}",
        ])
    if len(data_det) == 1:
        data_det.append(["Sin movimientos", "-", "-", "-", "-", "-"])

    table_det = Table(data_det, colWidths=[80, 110, 70, 75, 55, 75])
    style_det = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
    ]
    for row_idx in range(1, len(data_det)):
        if row_idx % 2 == 1:
            style_det.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor('#f3f4f6')))
    table_det.setStyle(TableStyle(style_det))
    story.append(table_det)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _crear_nombre_archivo(tipo_reporte: str, formato: str) -> str:
    sufijo = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"reportes-{tipo_reporte}-{sufijo}.{formato}"


@app.get("/reportes/export/")
def exportar_reportes(tipo_reporte: str = "mensual", format: str = "xlsx", start_date: Optional[datetime] = None, end_date: Optional[datetime] = None, db: Session = Depends(get_db)):
    formato = format.lower()
    if formato not in ["xlsx", "pdf"]:
        raise HTTPException(status_code=400, detail="Formato de exportación no válido. Usa pdf o xlsx.")

    resultados = _query_reporte(db, tipo_reporte.lower(), start_date, end_date)
    pagos_detalle = _query_pagos_detalle(db, start_date, end_date)

    if formato == "xlsx":
        contenido = _crear_excel_reporte(resultados, pagos_detalle, tipo_reporte, start_date, end_date)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        contenido = _crear_pdf_reporte(resultados, pagos_detalle, tipo_reporte, start_date, end_date)
        media_type = "application/pdf"

    return StreamingResponse(
        io.BytesIO(contenido),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename=\"{_crear_nombre_archivo(tipo_reporte, formato)}\""}
    )


@app.get("/reportes/", response_model=ReporteResponse)
def obtener_reportes(tipo_reporte: str = "mensual", start_date: Optional[datetime] = None, end_date: Optional[datetime] = None, db: Session = Depends(get_db)):
    """Genera un reporte agregado por período y devuelve totales con rangos."""
    resultados = _query_reporte(db, tipo_reporte.lower(), start_date, end_date)
    totales = _agregar_total_reporte(resultados)
    return {
        "tipo_reporte": tipo_reporte,
        "resultados": resultados,
        **totales,
    }

# --- 1. RUTA PARA SUBIR IMAGEN (Corrección para que funcione el botón) ---
@app.post("/subir-pago/")
async def subir_pago(
    file: UploadFile = File(...),
    banco: str = Form(...),
    cliente_id: Optional[str] = Form(None),
    comentario: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    # --- MEJORA: Seguridad con API Key ---
    auth: bool = Depends(require_api_key)
):
    # --- MEJORA: Endpoint mucho más robusto ---
    filepath = None
    try:
        if cliente_id == "":
            cliente_id = None
        elif cliente_id is not None:
            try:
                cliente_id = int(cliente_id)
            except ValueError:
                raise HTTPException(status_code=422, detail="cliente_id debe ser un entero válido o vacío.")
        # 1. Validaciones y guardado seguro
        if not file.content_type or not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="Solo se permiten archivos de imagen.")

        max_upload_mb = int(os.getenv("MAX_UPLOAD_MB", "10"))
        max_bytes = max_upload_mb * 1024 * 1024

        # Nombre de archivo sanitizado
        filename = f"{uuid.uuid4().hex}_{os.path.basename(file.filename)}"
        filepath = str(uploads_dir / filename)

        # Guardado en streaming con límite de tamaño
        written_bytes = 0
        with open(filepath, "wb") as buffer:
            while chunk := await file.read(1024 * 1024): # Leemos en bloques de 1MB
                written_bytes += len(chunk)
                if written_bytes > max_bytes:
                    raise HTTPException(status_code=413, detail=f"Archivo demasiado grande (máx {max_upload_mb}MB)")
                buffer.write(chunk)

        # 2. Calcular Hash para anti-duplicados
        sha256_hash = hashlib.sha256()
        with open(filepath, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        file_hash_str = sha256_hash.hexdigest()

        # --- MEJORA: Validación anti-duplicado por HASH más informativa ---
        existing_by_hash = db.query(models.Pago).filter(models.Pago.file_hash == file_hash_str).first()
        if existing_by_hash:
            return JSONResponse(
                content={
                    "mensaje": "Archivo duplicado detectado (hash). Ya existe un pago con este archivo.", 
                    "id_existente": existing_by_hash.id, 
                    "referencia": existing_by_hash.referencia
                },
                status_code=409
            )

        # 3. Procesar con OCR
        try:
            resultado = ocr_engine.procesar_imagen(filepath)
        except RuntimeError as e:
            logger.error("Error crítico OCR: %s", e)
            raise HTTPException(status_code=500, detail=str(e))

        # 4. Validación anti-duplicado por lógica de negocio (se mantiene tu lógica)
        # Priorizamos los datos limpios devueltos por la IA (Groq)
        ref_ocr = str(resultado.get("referencia", "No detectada"))
        
        try:
            # Aseguramos que el monto sea float para cálculos y base de datos
            monto_ocr = float(resultado.get("monto") or 0.0)
        except (ValueError, TypeError):
            monto_ocr = 0.0

        if ref_ocr not in ["S/R", "No detectada"]:
            pago_existente = db.query(models.Pago).filter(
                models.Pago.referencia == ref_ocr,
                models.Pago.banco == banco,
                models.Pago.monto == monto_ocr
            ).first()
            if pago_existente:
                return JSONResponse(
                    content={"mensaje": f"Error: Ya existe un pago con la misma referencia ({ref_ocr}), banco ({banco}) and monto."}, 
                    status_code=409
                )

        # 5. Guardar en BD
        banco_destino = resultado.get("banco_destino") # Puede ser None

        # Obtener la tasa más reciente con la nueva ruta de intercambio
        tasa_info = await get_tasa_bcv(db)
        tasa_actual = float(tasa_info['tasa'])
        monto_usd_calc = round(monto_ocr / tasa_actual, 2) if tasa_actual > 0 else 0.0

        nuevo_pago = models.Pago(
            referencia=ref_ocr,
            banco=banco,
            banco_destino=banco_destino,
            monto=monto_ocr,
            ruta_imagen=filepath,
            file_hash=file_hash_str,
            cliente_id=cliente_id,
            tasa_momento=tasa_actual,
            tasa_cambio=tasa_actual,
            monto_usd=monto_usd_calc
        )
        
        try:
            db.add(nuevo_pago)
            db.commit()
            db.refresh(nuevo_pago)
        except sa_exc.IntegrityError: # --- MEJORA: Manejo de carrera (dos uploads idénticos a la vez)
            db.rollback()
            existing = db.query(models.Pago).filter(models.Pago.file_hash == file_hash_str).first()
            if existing:
                return JSONResponse(
                    content={
                        "mensaje": "Archivo duplicado detectado (hash). Ya existe un pago con este archivo.",
                        "id_existente": existing.id,
                        "referencia": existing.referencia,
                    },
                    status_code=409
                )
            # Si no es por el hash, relanzamos el error
            raise


        # 6. Auditoría (PagoHistory) - Se mantiene tu lógica
        origen_datos = resultado.get("source", "RULES_LEGACY") # Default si no viene
        
        detalles_audit = f"Creado vía /subir-pago/. Motor: {origen_datos}. Ref: {ref_ocr}. Banco: {nuevo_pago.banco}"
        if banco_destino:
             detalles_audit += f" -> Destino: {banco_destino}"

        registrar_auditoria(
            db, 
            pago_id=nuevo_pago.id, 
            accion="create_ia", 
            detalles=detalles_audit
        )
        
        return JSONResponse(content={"mensaje": "Procesado correctamente", "data": resultado}, status_code=200)

    except HTTPException:
        raise # Re-lanza las excepciones HTTP para que FastAPI las maneje
    except Exception as e:
        # --- MEJORA: Logging de errores detallado sin exponerlo al cliente ---
        tb = traceback.format_exc()
        logger.error("Fallo crítico en /subir-pago/\n%s", tb)
        try:
            with open('logs/subir_trace.log', 'a', encoding='utf-8') as fh:
                fh.write(f"--- {datetime.now()} ---\n{tb}\n")
        except Exception as log_e:
            logger.error("No se pudo escribir el log de trace: %s", log_e)
        raise HTTPException(status_code=500, detail="Error interno del servidor al procesar el archivo.")
    finally:
        if filepath and os.path.exists(filepath) and ('nuevo_pago' not in locals() or not getattr(locals().get('nuevo_pago'), 'id', None)):
            os.remove(filepath)

@app.post("/detectar-banco/")
async def detectar_banco(
    file: UploadFile = File(...),
    auth: bool = Depends(require_api_key)
):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos de imagen.")

    temp_file = tempfile.NamedTemporaryFile(suffix=os.path.splitext(file.filename)[1] or ".jpg", delete=False)
    temp_path = temp_file.name
    temp_file.close()
    try:
        with open(temp_path, "wb") as buffer:
            while chunk := await file.read(1024 * 1024):
                buffer.write(chunk)

        resultado = ocr_engine.procesar_imagen(temp_path)
        if not resultado:
            raise HTTPException(status_code=500, detail="No se pudo procesar la imagen para detección de banco.")

        return {
            "banco_predicho": resultado.get("banco_predicho"),
            "banco_ia": resultado.get("banco_ia"),
            "sudeban_code": resultado.get("sudeban_code"),
            "referencia": resultado.get("referencia"),
            "monto": resultado.get("monto"),
            "cedula": resultado.get("cedula"),
            "texto_completo": resultado.get("texto_completo"),
        }
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass

@app.post("/detectar-banco-vision/")
async def detectar_banco_vision(
    data: VisionBankDetectionRequest,
    auth: bool = Depends(require_api_key)
):
    image_base64 = data.image_base64.strip()
    if not image_base64:
        raise HTTPException(status_code=400, detail="Se requiere image_base64.")

    try:
        image_data = base64.b64decode(image_base64.split(",", 1)[-1])
    except Exception:
        raise HTTPException(status_code=400, detail="Base64 inválido para la imagen.")

    temp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    temp_path = temp_file.name
    temp_file.close()
    try:
        with open(temp_path, "wb") as buffer:
            buffer.write(image_data)

        resultado = ocr_engine.procesar_imagen(temp_path) or {}
        groq_info = _detectar_banco_con_groq(image_data)

        banco_predicho = groq_info.get("banco_predicho") or resultado.get("banco_predicho") or resultado.get("banco_ia")
        sudeban_code = groq_info.get("sudeban_code") or resultado.get("sudeban_code")

        return {
            "banco_predicho": banco_predicho,
            "banco_ia": resultado.get("banco_ia"),
            "sudeban_code": sudeban_code,
            "referencia": resultado.get("referencia"),
            "monto": resultado.get("monto"),
            "cedula": resultado.get("cedula"),
            "texto_completo": resultado.get("texto_completo"),
        }
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


def _parse_groq_bank_response(texto: str) -> dict:
    if not texto:
        return {}
    texto = texto.strip()
    try:
        encontrado = re.search(r"\{.*\}", texto, re.S)
        if encontrado:
            return json.loads(encontrado.group(0))
    except Exception:
        pass

    parsed = {}
    match = re.search(r'"?banco_predicho"?\s*[:=]\s*"?([^"\'\n]+)', texto, re.I)
    if match:
        parsed["banco_predicho"] = match.group(1).strip()
    match = re.search(r'"?sudeban_code"?\s*[:=]\s*"?([^"\'\n]+)', texto, re.I)
    if match:
        parsed["sudeban_code"] = match.group(1).strip()
    if not parsed:
        parsed["banco_predicho"] = texto.splitlines()[0].strip()
    return parsed


def _comprimir_imagen_para_groq(image_bytes: bytes, max_side: int = 720, quality: int = 60) -> bytes:
    try:
        with Image.open(io.BytesIO(image_bytes)) as img:
            img = img.convert("RGB")
            max_dimension = max(img.width, img.height)
            if max_dimension > max_side:
                scale = max_side / max_dimension
                img = img.resize(
                    (max(1, int(img.width * scale)), max(1, int(img.height * scale))),
                    Image.Resampling.LANCZOS,
                )
            with io.BytesIO() as output:
                img.save(output, format="JPEG", quality=quality, optimize=True)
                return output.getvalue()
    except Exception as e:
        logger.debug("No se pudo comprimir la imagen para Groq, se usa el original: %s", e)
        return image_bytes


def _detectar_banco_con_groq(image_bytes: bytes) -> dict:
    api_key = os.getenv("GROQ_API_KEY", "").strip()
    if not api_key:
        return {}

    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        image_bytes_for_groq = _comprimir_imagen_para_groq(image_bytes)
        image_b64 = base64.b64encode(image_bytes_for_groq).decode("utf-8")
        prompt_text = (
            "Eres un experto en reconocer bancos venezolanos a partir de comprobantes de pago. "
            "Devuelve un JSON válido con los campos banco_predicho y sudeban_code. "
            "Si no puedes identificar el banco, usa Desconocido. "
            "Responde únicamente con JSON válido, sin texto adicional."
        )
        response = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": image_b64}},
                        {"type": "text", "text": prompt_text},
                    ],
                }
            ],
            model=os.getenv("GROQ_MODEL", "llama-3.2-11b-vision-preview"),
            temperature=0.0,
            max_tokens=150,
        )
        content = ""
        if getattr(response, 'choices', None):
            choice = response.choices[0]
            message = getattr(choice, 'message', None)
            if isinstance(message, dict):
                content = message.get('content', '')
            elif hasattr(message, 'content'):
                content = message.content or ''
            else:
                content = str(message or '')
        return _parse_groq_bank_response(content)
    except Exception as e:
        logger.warning("Groq Vision fallo: %s", e)
        return {}


# --- 2. RUTA PARA REGISTRO MANUAL (Nueva) ---
@app.post("/pago-manual/")
async def crear_pago_manual(dato: PagoManual, db: Session = Depends(get_db)):
    try:
        if not dato.banco.strip():
            raise HTTPException(status_code=400, detail="El banco es obligatorio y debe ser un valor válido.")
        if not dato.referencia.strip():
            raise HTTPException(status_code=400, detail="La referencia es obligatoria y no puede estar vacía.")
        if dato.monto <= 0:
            raise HTTPException(status_code=400, detail="El monto debe ser un número mayor a cero.")

        # --- VALIDACIÓN ANTI-DUPLICADO MEJORADA ---
        pago_existente = db.query(models.Pago).filter(
            models.Pago.referencia == dato.referencia,
            models.Pago.banco == dato.banco,
            models.Pago.monto == dato.monto
        ).first()
        if pago_existente:
            raise HTTPException(status_code=409, detail=f"Error: Ya existe un pago con la misma referencia, banco y monto.")

        # Normalizar banco manual contra la lista canónica de bank_rules
        banco_normalizado = dato.banco.strip()
        if banco_normalizado not in bank_rules.get_available_banks():
            try:
                estrategia = bank_rules.get_bank_strategy(banco_normalizado)
                if estrategia and estrategia.name and estrategia.name != "Desconocido":
                    banco_normalizado = estrategia.name
            except Exception:
                pass

        # Obtener tasa para el registro manual (NCBC nueva función)
        tasa_info = await get_tasa_bcv(db)
        tasa_actual = float(tasa_info["tasa"])
        monto_usd_calc = round(dato.monto / tasa_actual, 2) if tasa_actual > 0 else 0.0

        nuevo_pago = models.Pago(
            referencia=dato.referencia,
            banco=banco_normalizado,
            monto=dato.monto,
            ruta_imagen="", # Vacío porque es manual
            file_hash=None,
            cliente_id=dato.cliente_id,
            tasa_momento=tasa_actual,
            tasa_cambio=tasa_actual,
            monto_usd=monto_usd_calc
        )
        db.add(nuevo_pago)
        db.commit()
        db.refresh(nuevo_pago)

        registrar_auditoria(
            db, 
            pago_id=nuevo_pago.id, 
            accion="create_manual", 
            detalles="Registro manual por el usuario"
        )
        return {"mensaje": "Pago manual registrado"}
    except Exception as e:
        logger.error(f"Error manual: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.patch("/pagos/{pago_id}", response_model=PagoResponse)
async def actualizar_pago_manual(pago_id: int, datos: PagoManual, db: Session = Depends(get_db)):
    pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    if pago.ruta_imagen:
        raise HTTPException(status_code=400, detail="Solo se pueden editar pagos manuales sin imagen.")

    if not datos.banco or not datos.referencia or datos.monto is None:
        raise HTTPException(
            status_code=400,
            detail="Banco, referencia y monto son obligatorios. Asegúrate de completar todos los campos y que el monto sea mayor a cero."
        )

    pago.banco = datos.banco.strip()
    pago.referencia = datos.referencia.strip()
    pago.monto = datos.monto
    pago.cliente_id = datos.cliente_id

    tasa_info = await get_tasa_bcv(db)
    nueva_tasa = float(tasa_info["tasa"])
    pago.tasa_momento = nueva_tasa
    pago.tasa_cambio = nueva_tasa
    pago.monto_usd = round(datos.monto / nueva_tasa, 2) if nueva_tasa > 0 else 0.0

    db.commit()
    db.refresh(pago)

    registrar_auditoria(
        db,
        pago_id=pago.id,
        accion="update_manual",
        detalles=f"Pago manual editado: ref {pago.referencia}, monto {pago.monto}, banco {pago.banco}"
    )
    return pago

# --- 3. RUTA PARA RE-PROCESAR (Nueva) ---
@app.post("/reprocesar/{pago_id}")
def reprocesar_pago(pago_id: int, db: Session = Depends(get_db)):
    # 1. Buscar el pago
    pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    
    if not pago.ruta_imagen or not os.path.exists(pago.ruta_imagen):
        raise HTTPException(status_code=400, detail="Este pago no tiene imagen para reprocesar")

    try:
        # 2. Ejecutar OCR de nuevo
        resultado = ocr_engine.procesar_imagen(pago.ruta_imagen)
        
        # 3. Actualizar datos (No borramos para mantener el ID, solo actualizamos)
        banco_dest_nuevo = resultado.get("banco_destino", pago.banco_destino)
        ref_nueva = resultado.get("referencia", pago.referencia)
        monto_nuevo = float(resultado.get("monto") or pago.monto)

        pago.referencia = ref_nueva
        pago.banco_destino = banco_dest_nuevo
        pago.monto = monto_nuevo
        
        db.commit()

        # Mejoramos el detalle de la auditoría
        detalles_audit = f"Re-escaneo IA. Ref: {ref_nueva}, Banco: {pago.banco}, Monto: {monto_nuevo}"

        registrar_auditoria(
            db, 
            pago_id=pago.id, 
            accion="reprocess", 
            detalles=detalles_audit
        )
        return {"mensaje": "Pago reprocesado con éxito", "nuevos_datos": resultado}
    except Exception as e:
        logger.error(f"Error reprocesando: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# --- NUEVA RUTA PARA CAMBIAR ESTADO ---
@app.patch("/pago/{pago_id}/estado")
def cambiar_estado_pago(pago_id: int, update_data: EstadoUpdate, db: Session = Depends(get_db)):
    """Actualiza el estado de un pago y registra en la auditoría."""
    pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    estado_anterior = pago.estado
    nuevo_estado = update_data.estado.value
    
    if estado_anterior == nuevo_estado:
        return {"mensaje": "El estado ya es el solicitado, no se realizaron cambios."}

    pago.estado = nuevo_estado
    
    registrar_auditoria(
        db,
        pago_id=pago.id,
        accion="update_status",
        detalles=f"Estado cambiado de '{estado_anterior}' a '{nuevo_estado}'"
    )
    return {"mensaje": f"Estado del pago {pago_id} actualizado a '{nuevo_estado}'"}

@app.get("/pago/{pago_id}/historial")
def obtener_historial(pago_id: int, db: Session = Depends(get_db)):
    # Buscamos todos los registros de auditoría para ese pago
    historial = db.query(models.PagoHistory).filter(
        models.PagoHistory.pago_id == pago_id
    ).order_by(desc(models.PagoHistory.id)).all()
    return historial

@app.get("/pagos/{pago_id}/imagen")
def obtener_imagen_pago(pago_id: int, db: Session = Depends(get_db)):
    pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    if not pago.ruta_imagen:
        raise HTTPException(status_code=404, detail="Este pago no tiene imagen asociada")

    filename = os.path.basename(pago.ruta_imagen)
    if not filename:
        raise HTTPException(status_code=404, detail="Ruta de imagen inválida")

    return {"imagen_url": f"/uploads/{filename}"}

@app.post("/pagos/{pago_id}/reprocesar")
def reprocesar_pago_alias(pago_id: int, db: Session = Depends(get_db)):
    return reprocesar_pago(pago_id, db)

@app.post("/pagos/{pago_id}/estado")
def cambiar_estado_pago_alias(pago_id: int, update_data: EstadoUpdate, db: Session = Depends(get_db)):
    return cambiar_estado_pago(pago_id, update_data, db)

@app.delete("/pagos/{pago_id}/")
def eliminar_pago_por_id(pago_id: int, db: Session = Depends(get_db)):
    pago = db.query(models.Pago).filter(models.Pago.id == pago_id).first()
    if not pago:
        raise HTTPException(status_code=404, detail="Pago no encontrado")

    if pago.ruta_imagen and os.path.exists(pago.ruta_imagen):
        try:
            os.remove(pago.ruta_imagen)
        except Exception as e:
            logger.warning(f"No se pudo eliminar el archivo físico {pago.ruta_imagen}: {e}")

    db.delete(pago)
    db.commit()
    return {"mensaje": "Eliminado correctamente"}

@app.delete("/eliminar-pago-ref/{referencia}")
def eliminar_pago(referencia: str, confirm: bool = False, db: Session = Depends(get_db)):
    pago = db.query(models.Pago).filter(models.Pago.referencia == referencia).first()
    if not pago:
        # Intentamos buscar por ID si la referencia parece un ID (fallback)
        if referencia.isdigit():
             pago = db.query(models.Pago).filter(models.Pago.id == int(referencia)).first()
        
        if not pago:
            raise HTTPException(status_code=404, detail="Pago no encontrado")

    if confirm:
        # Borrar imagen física si existe
        if pago.ruta_imagen and os.path.exists(pago.ruta_imagen):
            try:
                os.remove(pago.ruta_imagen)
            except Exception as e:
                logger.warning(f"No se pudo eliminar el archivo físico {pago.ruta_imagen}: {e}")
        db.delete(pago)
        db.commit()
        return {"mensaje": "Eliminado correctamente"}
    
    return {"mensaje": "Se requiere confirmación (?confirm=true)"}